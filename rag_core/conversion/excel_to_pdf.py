"""
Excel to PDF conversion utilities for RAG API (UNO-free, headless)
- Fits each sheet to 1 page wide x 1 page tall
- Landscape orientation
- Attempts to include images/charts by expanding print area to their anchors
- Optional conservative fallback: print entire sheet if unparsed drawings exist
"""
import logging
import os
import shutil
import subprocess
import sys
import tempfile
import time
from pathlib import Path
from typing import Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from rag_core.config import config
from .lool_client import LoolClient

logger = logging.getLogger(__name__)

EXCEL_EXTENSIONS = (".xlsx", ".xls")
OFFICE_EXTENSIONS = EXCEL_EXTENSIONS + (".docx", ".doc", ".pptx", ".ppt")

_LATEST_DEBUG_ARTIFACTS: Optional[Path] = None
_LATEST_DEBUG_COMBINED_PDF: Optional[Path] = None

# ---------- Config ----------
MAC_SOFFICE = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
INCLUDE_FULL_SHEET_IF_UNKNOWN_DRAWINGS = True   # safest to avoid clipping SmartArt
SHEET_MAX_COL = 16384  # XFD
SHEET_MAX_ROW = 1048576
# ---------------------------

def _soffice_bin() -> str:
    # Allow override via env; otherwise pick a sensible default
    override = os.environ.get("SOFFICE_BIN")
    if override:
        return override
    if sys.platform == "darwin" and Path(MAC_SOFFICE).exists():
        return MAC_SOFFICE
    return "soffice"

def _run(cmd: list, timeout: int = 600) -> subprocess.CompletedProcess:
    start = time.time()
    logger.debug("[SOFFICE] Running: %s", " ".join(cmd))
    proc = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    elapsed = time.time() - start
    logger.debug("[SOFFICE] rc=%s elapsed=%.2fs\nSTDOUT:\n%s\nSTDERR:\n%s",
                 proc.returncode, elapsed, proc.stdout.strip(), proc.stderr.strip())
    if proc.returncode != 0:
        raise RuntimeError(
            f"Command failed ({proc.returncode}): {' '.join(cmd)}\n"
            f"STDOUT:\n{proc.stdout}\nSTDERR:\n{proc.stderr}"
        )
    return proc

def _format_a1(min_col: int, min_row: int, max_col: int, max_row: int) -> str:
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

def _cell_used_bounds(ws):
    # openpyxl's calculate_dimension returns like "A1:D42" or "A1:A1" if empty
    dim = ws.calculate_dimension()
    try:
        min_col, min_row, max_col, max_row = range_boundaries(dim)
        return min_col, min_row, max_col, max_row
    except Exception:
        # Fallback: at least include A1
        return 1, 1, 1, 1

def _anchor_bounds(anchor):
    """
    Try to extract bounds from image/chart anchor.
    Supports: "A1" style, TwoCellAnchor, OneCellAnchor best-effort.
    Returns (min_col, min_row, max_col, max_row) or None.
    """
    try:
        # Simple "A1" style
        if isinstance(anchor, str):
            c = anchor
            # "A1" -> treat as 1x1
            col = 1
            row = 1
            # safer parse:
            from openpyxl.utils.cell import coordinate_to_tuple
            row, col = coordinate_to_tuple(c)  # note: returns (row, col)
            return col, row, col, row

        # TwoCellAnchor: has _from and _to with .col/.row (0-based)
        if hasattr(anchor, "_from") and hasattr(anchor, "_to"):
            fc = getattr(anchor._from, "col", 0)
            fr = getattr(anchor._from, "row", 0)
            tc = getattr(anchor._to, "col", fc)
            tr = getattr(anchor._to, "row", fr)
            # convert to 1-based cell indices
            min_col = min(fc, tc) + 1
            max_col = max(fc, tc) + 1
            min_row = min(fr, tr) + 1
            max_row = max(fr, tr) + 1
            return min_col, min_row, max_col, max_row

        # OneCellAnchor: has _from plus extents in EMUs; approximate to one cell
        if hasattr(anchor, "_from"):
            fc = getattr(anchor._from, "col", 0)
            fr = getattr(anchor._from, "row", 0)
            return fc + 1, fr + 1, fc + 1, fr + 1

    except Exception:
        pass
    return None

def _merge_bounds(a, b):
    if not a: return b
    if not b: return a
    min_col = min(a[0], b[0]); min_row = min(a[1], b[1])
    max_col = max(a[2], b[2]); max_row = max(a[3], b[3])
    return (min_col, min_row, max_col, max_row)

def _sheet_has_unknown_drawings(ws) -> bool:
    """
    If there is a drawing container but we didn't see images/charts,
    assume there might be SmartArt or other shapes we can't bound.
    """
    try:
        # openpyxl keeps a private _drawing ref if there are any drawings
        return getattr(ws, "_drawing", None) is not None
    except Exception:
        return False

def _expand_bounds_with_drawings(ws, bounds):
    """
    Attempt to include images and charts in the print area.
    If unknown drawings exist and flag is enabled, return full-sheet bounds.
    """
    acc = bounds
    saw_any_known = False

    # Images (private attr, but widely used)
    imgs = getattr(ws, "_images", []) or []
    for img in imgs:
        b = _anchor_bounds(getattr(img, "anchor", None))
        if b:
            acc = _merge_bounds(acc, b)
            saw_any_known = True

    # Charts (private attr)
    charts = getattr(ws, "_charts", []) or []
    for ch in charts:
        b = _anchor_bounds(getattr(ch, "anchor", None))
        if b:
            acc = _merge_bounds(acc, b)
            saw_any_known = True

    if INCLUDE_FULL_SHEET_IF_UNKNOWN_DRAWINGS:
        if _sheet_has_unknown_drawings(ws) and not saw_any_known:
            # Conservative: include the entire grid so SmartArt/shapes are not clipped
            return (1, 1, SHEET_MAX_COL, SHEET_MAX_ROW)

    return acc

def _patch_pagesetup_xlsx(src_xlsx: Path) -> Path:
    """
    Create a temp .xlsx that forces:
      - Landscape
      - Fit to 1 x 1 pages
      - Print area expanded to include images/charts (and optionally all drawings)
    """
    tmp_xlsx = Path(tempfile.mkstemp(suffix=".xlsx")[1])
    shutil.copy2(src_xlsx, tmp_xlsx)  # do not mutate original

    wb = load_workbook(tmp_xlsx)
    for ws in wb.worksheets:
        # Base bounds from used cells
        base = _cell_used_bounds(ws)
        # Expand with images/charts; optionally full sheet if unknown drawings exist
        final = _expand_bounds_with_drawings(ws, base)

        # Page setup: force landscape + fit-to-page
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        # Avoid explicit scale if present
        try:
            ws.page_setup.scale = None
        except Exception:
            pass

        # Margins: keep defaults; optionally shrink to maximize content
        # ws.page_margins.left = ws.page_margins.right = 0.25
        # ws.page_margins.top = ws.page_margins.bottom = 0.5

        # Print area
        if final:
            min_c, min_r, max_c, max_r = final
            # Guard against insane bounds (stay within sheet limits)
            min_c = max(1, min_c); min_r = max(1, min_r)
            max_c = min(SHEET_MAX_COL, max_c); max_r = min(SHEET_MAX_ROW, max_r)
            ws.print_area = _format_a1(min_c, min_r, max_c, max_r)

    wb.save(tmp_xlsx)
    return tmp_xlsx

def _convert_with_soffice(input_path: Path, outdir: Path, filter_str: str) -> Path:
    soffice = _soffice_bin()
    outdir.mkdir(parents=True, exist_ok=True)
    cmd = [soffice, "--headless", "--convert-to", filter_str, "--outdir", str(outdir), str(input_path)]
    _run(cmd)

    # LibreOffice writes <stem>.<ext> (usually). Resolve robustly:
    produced = outdir / f"{input_path.stem}.{filter_str.split(':', 1)[0]}"
    if not produced.exists():
        # scan outdir as fallback
        candidates = sorted(outdir.glob(f"{input_path.stem}.*"), key=lambda p: p.stat().st_mtime, reverse=True)
        for c in candidates:
            if c.suffix.lower() in (".pdf", ".xlsx", ".xls"):
                produced = c
                break
    if not produced.exists():
        raise RuntimeError(f"[SOFFICE] Did not produce output for {input_path}")
    return produced

def _xls_to_xlsx_via_soffice(xls_path: Path, outdir: Path) -> Path:
    xlsx = _convert_with_soffice(xls_path, outdir, "xlsx")
    if xlsx.suffix.lower() != ".xlsx":
        raise RuntimeError(f"[SOFFICE] Unexpected output for {xls_path}: {xlsx}")
    return xlsx


def _maybe_generate_debug_artifacts(input_path: Path) -> Tuple[Optional[Path], Optional[Path]]:
    """Optionally keep split Excel/PDF artifacts for debugging."""
    try:
        from Convert_excel_pdf import config as convert_config
        from Convert_excel_pdf.excel_splitter import split_xlsx
        from Convert_excel_pdf.convert_parts_combine import convert_sheets_and_combine
    except Exception as exc:
        logger.warning(
            "[DEBUG] Skipping debug artifact generation for %s (imports unavailable): %s",
            input_path,
            exc,
        )
        return None, None

    if not getattr(convert_config, "GENERATE_DEBUG_ARTIFACTS", True):
        logger.info("[DEBUG] Debug artifact generation disabled via Convert_excel_pdf config")
        return None, None

    if input_path.suffix.lower() != ".xlsx":
        logger.info("[DEBUG] Debug artifacts currently generated for .xlsx files only: %s", input_path)
        return None, None

    base_dir = convert_config.SPLIT_OUTPUT_ROOT / input_path.stem
    if base_dir.exists():
        shutil.rmtree(base_dir)

    logger.info(
        "[DEBUG] Generating debug artifacts for %s under %s (max sheets %s)",
        input_path,
        base_dir,
        getattr(convert_config, "MAX_SHEETS_PER_SPLIT", None),
    )

    try:
        outputs = split_xlsx(
            input_path,
            convert_config.SPLIT_OUTPUT_ROOT,
            convert_config.MAX_SHEETS_PER_SPLIT,
        )
        if not outputs:
            logger.warning("[DEBUG] No split Excel outputs created for %s", input_path)
            return None, None

        pdf_dir = base_dir / "pdf"
        combined_pdf = pdf_dir / f"combined_{input_path.stem}.pdf"
        convert_sheets_and_combine(
            sheet_excels=outputs,
            pdf_dir=pdf_dir,
            combined_pdf=combined_pdf,
            lool_url=config.LOOL_BASE_URL,
            endpoint=config.LOOL_ENDPOINT,
            timeout=config.LOOL_TIMEOUT,
        )
        logger.info(
            "[DEBUG] Generated debug artifacts under %s (sheets + PDFs)",
            base_dir,
        )
        global _LATEST_DEBUG_ARTIFACTS, _LATEST_DEBUG_COMBINED_PDF
        _LATEST_DEBUG_ARTIFACTS = base_dir
        _LATEST_DEBUG_COMBINED_PDF = combined_pdf
        return base_dir, combined_pdf
    except Exception as exc:
        logger.warning("[DEBUG] Failed to generate debug artifacts for %s: %s", input_path, exc)
        return None, None


def _convert_excel_with_lool(input_path: Path, output_dir: Path) -> Path:
    """Convert Excel workbook via LibreOffice Online FullSheetPreview API."""
    output_dir.mkdir(parents=True, exist_ok=True)
    client = LoolClient(
        base_url=config.LOOL_BASE_URL,
        endpoint_path=config.LOOL_ENDPOINT,
        timeout=config.LOOL_TIMEOUT,
        retry_attempts=config.LOOL_RETRY_ATTEMPTS,
        retry_delay=config.LOOL_RETRY_DELAY,
    )

    logger.info("[CONVERSION] Sending workbook to LOOL: %s", input_path)
    pdf_bytes = client.convert_all_sheets(input_path)
    pdf_path = output_dir / f"{input_path.stem}.pdf"
    pdf_path.write_bytes(pdf_bytes)
    logger.info(
        "[CONVERSION] LOOL conversion succeeded: %s -> %s (%.1f KB)",
        input_path,
        pdf_path,
        pdf_path.stat().st_size / 1024,
    )
    return pdf_path

def _excel_to_pdf_no_uno(input_path: Path, office_outdir: Path) -> Path:
    """
    Excel (.xlsx/.xls) -> PDF (UNO-free):
      - For .xls: convert to .xlsx first
      - Patch page setup (landscape + 1x1 + print area incl. drawings)
      - Export to PDF via calc_pdf_Export
    """
    t0 = time.time()
    logger.info("[CONVERSION] Processing Excel file (UNO-free): %s", input_path)

    working_dir = office_outdir / "excel2pdf_tmp"
    working_dir.mkdir(parents=True, exist_ok=True)

    if input_path.suffix.lower() == ".xls":
        logger.debug("[CONVERSION] .xls detected; converting to .xlsx first")
        xlsx_path = _xls_to_xlsx_via_soffice(input_path, working_dir)
    else:
        xlsx_path = input_path

    logger.debug("[PATCH] Patching page setup for: %s", xlsx_path)
    patched = _patch_pagesetup_xlsx(Path(xlsx_path))

    logger.debug("[EXPORT] Exporting patched workbook to PDF")
    pdf_path = _convert_with_soffice(patched, office_outdir, "pdf:calc_pdf_Export")

    logger.info("[CONVERSION] Excel conversion successful: %s → %s (%.1f KB) in %.2fs",
                input_path, pdf_path, pdf_path.stat().st_size/1024, time.time() - t0)
    return pdf_path


def _convert_excel_document(input_path: Path, output_dir: Path) -> Path:
    """
    Convert an Excel workbook using LOOL when enabled, falling back to the
    local soffice-based pipeline if necessary.
    """
    last_exc: Optional[Exception] = None
    pdf_path: Optional[Path] = None

    if config.LOOL_ENABLED:
        try:
            pdf_path = _convert_excel_with_lool(input_path, output_dir)
        except Exception as exc:
            last_exc = exc
            logger.error(
                "[CONVERSION] LOOL conversion failed for %s: %s",
                input_path,
                exc,
            )
            if not config.LOOL_FALLBACK_TO_LOCAL:
                raise
            logger.info(
                "[CONVERSION] Falling back to local soffice pipeline for %s",
                input_path,
            )

    if pdf_path is None:
        try:
            pdf_path = _excel_to_pdf_no_uno(input_path, output_dir)
        except Exception as exc:
            if last_exc:
                logger.error(
                    "[CONVERSION] Local fallback after LOOL failure also failed for %s: %s",
                    input_path,
                    exc,
                )
            raise

    debug_dir, combined_pdf = _maybe_generate_debug_artifacts(input_path)
    if combined_pdf and combined_pdf.exists():
        target_path = output_dir / f"{input_path.stem}.pdf"
        try:
            shutil.copy2(combined_pdf, target_path)
            pdf_path = target_path
            logger.info(
                "[DEBUG] Replaced primary PDF with combined debug PDF: %s",
                target_path,
            )
        except Exception as copy_exc:
            logger.warning(
                "[DEBUG] Failed to copy combined debug PDF %s to %s: %s",
                combined_pdf,
                target_path,
                copy_exc,
            )
    if debug_dir:
        logger.info("[DEBUG] Debug artifacts available at %s", debug_dir)

    return pdf_path


def get_latest_debug_artifacts_dir() -> Optional[str]:
    """Return the most recent debug artifact directory generated."""
    if _LATEST_DEBUG_ARTIFACTS:
        return str(_LATEST_DEBUG_ARTIFACTS)
    return None


def get_latest_debug_combined_pdf() -> Optional[str]:
    """Return the most recent combined debug PDF path."""
    if _LATEST_DEBUG_COMBINED_PDF:
        return str(_LATEST_DEBUG_COMBINED_PDF)
    return None

def convert_office_to_pdf(input_path: str, output_dir: Optional[str] = None) -> Tuple[str, bool]:
    """Convert Office documents to PDF with comprehensive logging and timing (UNO-free for Excel)"""
    conversion_start = time.time()
    input_path = Path(input_path)
    ext = input_path.suffix.lower()

    logger.info("[CONVERSION] Starting Office→PDF conversion for: %s (extension: %s)", input_path, ext)

    if ext not in OFFICE_EXTENSIONS:
        logger.debug("[CONVERSION] File extension %s not in supported Office extensions, skipping conversion", ext)
        return str(input_path), False

    # Setup output directory
    if output_dir is None:
        temp_dir = tempfile.TemporaryDirectory()
        office_outdir = Path(temp_dir.name) / "office_pdf"
        logger.debug("[CONVERSION] Using temporary directory: %s", office_outdir)
    else:
        temp_dir = None
        office_outdir = Path(output_dir)
        logger.debug("[CONVERSION] Using specified output directory: %s", office_outdir)

    office_outdir.mkdir(parents=True, exist_ok=True)
    pdf_path = office_outdir / f"{input_path.stem}.pdf"

    try:
        if ext in EXCEL_EXTENSIONS:
            pdf_path = _convert_excel_document(input_path, office_outdir)
        else:
            # Standard conversion for Word/PowerPoint
            logger.info("[CONVERSION] Processing Office file with standard conversion: %s", input_path)
            conv = _convert_with_soffice(input_path, office_outdir, "pdf")
            pdf_path = conv

        total_elapsed = time.time() - conversion_start
        logger.info("[CONVERSION] Total conversion time: %.2fs", total_elapsed)
        return str(pdf_path), True

    except Exception as e:
        total_elapsed = time.time() - conversion_start
        logger.error("[CONVERSION] Office→PDF conversion failed after %.2fs for %s: %s",
                     total_elapsed, input_path, str(e))
        return str(input_path), False
    finally:
        if 'temp_dir' in locals() and temp_dir:
            temp_dir.cleanup()
            logger.debug("[CONVERSION] Cleaned up temporary directory")

def convert_excel_to_pdf(input_path: str, output_dir: Optional[str] = None) -> str:
    """
    Convert Excel file to PDF and save to input/converted directory.
    Ensures landscape + 1x1 pages; expands print area to include images/charts;
    optionally full-sheet if unparsed drawings (SmartArt) are present.
    """
    excel_start = time.time()
    input_path = Path(input_path)
    logger.info("[EXCEL] Excel conversion requested: %s", input_path)

    # Use input/converted as default output directory
    if output_dir is None:
        # Get the project root directory (assuming this file is in rag_core/conversion/)
        project_root = Path(__file__).parent.parent.parent
        output_dir = project_root / "input" / "converted"
        logger.debug("[EXCEL] Using default output directory: %s", output_dir)
    else:
        output_dir = Path(output_dir)
        logger.debug("[EXCEL] Using custom output directory: %s", output_dir)

    output_dir.mkdir(parents=True, exist_ok=True)

    ext = input_path.suffix.lower()
    if ext not in EXCEL_EXTENSIONS:
        logger.error("[EXCEL] Invalid file type: %s (expected .xlsx or .xls)", ext)
        raise ValueError(f"Input file {input_path} is not an Excel file")

    logger.debug("[EXCEL] Converting Excel file via configured pipeline")
    pdf_path = _convert_excel_document(input_path, output_dir)

    total_elapsed = time.time() - excel_start
    logger.info("[EXCEL] Excel conversion completed successfully in %.2fs: %s", total_elapsed, pdf_path)
    return str(pdf_path)

def needs_conversion(file_path: str) -> bool:
    """Check if a file needs to be converted to PDF."""
    ext = Path(file_path).suffix.lower()
    needs_conv = ext in OFFICE_EXTENSIONS
    logger.debug("[CONVERSION] File %s needs conversion: %s (extension: %s)", file_path, needs_conv, ext)
    return needs_conv
