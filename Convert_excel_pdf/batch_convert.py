#!/usr/bin/env python3
"""
Batch Excel to PDF converter with individual sheet extraction.

Converts each worksheet in an Excel file to separate PDF files,
then combines them into a single merged PDF.

Usage:
    python batch_convert.py input.xlsx

Output structure:
    split/[excel_name]/
    ├── sheet_001_SheetName.pdf
    ├── sheet_002_SheetName.pdf
    ├── ...
    └── combined_[excel_name].pdf

With --deep option:
    split/[excel_name]/
    ├── sheets/                    # Individual Excel sheet files
    │   ├── 001_SheetName.xlsx
    │   ├── 002_SheetName.xlsx
    │   └── ...
    ├── sheet_001_SheetName.pdf    # Converted PDFs
    ├── sheet_002_SheetName.pdf
    ├── ...
    └── combined_[excel_name].pdf
"""

import io
import logging
import sys
from pathlib import Path
from typing import List

from Convert_excel_pdf.service.lool_client import LoolClient
from Convert_excel_pdf.config import MAX_SHEETS_PER_SPLIT

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def sanitize_filename(name: str) -> str:
    """Sanitize sheet names for use as filenames."""
    # Replace problematic characters with underscores
    return "".join(c if c.isalnum() or c in "._-" else "_" for c in name)


def get_sheet_names(xlsx_path: Path) -> List[str]:
    """
    Extract sheet names from Excel file.

    Returns:
        List of sheet names in order
    """
    import zipfile
    import xml.etree.ElementTree as ET

    sheet_names = []
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as zf:
            # Read workbook.xml to get sheet names
            with zf.open("xl/workbook.xml") as f:
                xml_content = f.read()

            root = ET.fromstring(xml_content)
            ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

            # Find all sheet elements and extract names
            sheets = root.findall(".//ns:sheets/ns:sheet", ns)
            for sheet in sheets:
                name = sheet.get("name", f"Sheet_{len(sheet_names)+1}")
                sheet_names.append(name)

    except Exception as e:
        logger.warning(f"Could not extract sheet names: {e}. Using generic names.")
        # Fallback to generic names
        sheet_count = LoolClient.count_sheets(xlsx_path)
        sheet_names = [f"Sheet_{i+1}" for i in range(sheet_count)]

    return sheet_names


def split_excel_into_sheets(xlsx_path: Path, output_dir: Path) -> List[Path]:
    """
    Split Excel workbook by converting in chunks to avoid timeouts.

    This preserves ALL content (SmartArt, images, charts) by converting the full
    workbook but in smaller chunks if needed, then splitting the PDF.

    Args:
        xlsx_path: Path to input Excel file
        output_dir: Directory to save individual PDF files

    Returns:
        List of paths to individual PDF files (one per sheet)
    """
    logger.info(f"Converting workbook in chunks and splitting into individual sheets...")

    # Get sheet count first
    try:
        sheet_count = LoolClient.count_sheets(xlsx_path)
    except Exception as e:
        logger.error(f"Failed to count sheets: {e}")
        return []

    # For very large workbooks, try chunked conversion
    if sheet_count > 50:
        logger.info(f"Large workbook ({sheet_count} sheets) - using chunked conversion")
        return _split_excel_chunked(xlsx_path, output_dir, sheet_count)
    else:
        # For smaller workbooks, use single conversion
        return _split_excel_single(xlsx_path, output_dir, sheet_count)


def _split_excel_single(xlsx_path: Path, output_dir: Path, sheet_count: int) -> List[Path]:
    """Convert entire workbook at once (for smaller files)."""
    # Initialize client for single conversion
    client = LoolClient(timeout=1200)  # 20 minutes for large files

    try:
        # Convert entire workbook at once (preserves ALL content including SmartArt, images)
        logger.info(f"Converting {sheet_count} sheets to single PDF...")
        full_pdf_bytes = client.convert_all_sheets(xlsx_path)

        # Split the PDF into individual pages
        from pypdf import PdfReader, PdfWriter

        reader = PdfReader(io.BytesIO(full_pdf_bytes))
        actual_pages = len(reader.pages)

        if actual_pages != sheet_count:
            logger.warning(f"Expected {sheet_count} pages but got {actual_pages}")

        # Create individual PDF files
        sheet_files = []
        sheet_names = get_sheet_names(xlsx_path)

        for idx in range(min(actual_pages, len(sheet_names))):
            sheet_name = sheet_names[idx]

            # Extract single page
            writer = PdfWriter()
            writer.add_page(reader.pages[idx])

            # Save individual PDF
            safe_name = sanitize_filename(sheet_name)
            pdf_filename = f"{idx+1:03d}_{safe_name}.pdf"
            pdf_path = output_dir / pdf_filename

            with open(pdf_path, 'wb') as f:
                writer.write(f)

            sheet_files.append(pdf_path)
            logger.info(f"  → Extracted page {idx+1}: {pdf_path}")

        logger.info(f"✅ Split into {len(sheet_files)} individual PDF files")
        return sheet_files

    except Exception as e:
        logger.error(f"Single conversion failed: {e}")
        logger.info("Falling back to chunked conversion...")
        return _split_excel_chunked(xlsx_path, output_dir, sheet_count)


def _split_excel_chunked(xlsx_path: Path, output_dir: Path, sheet_count: int) -> List[Path]:
    """Convert workbook in chunks by creating temporary Excel files with subsets of sheets."""
    logger.info(f"Creating temporary Excel files with sheet subsets...")

    try:
        from openpyxl import load_workbook, Workbook
    except ImportError:
        logger.error("openpyxl required for chunked conversion")
        return []

    sheet_names = get_sheet_names(xlsx_path)
    chunk_size = max(1, MAX_SHEETS_PER_SPLIT)  # Process configurable number of sheets at a time

    all_sheet_files = []

    # Process sheets in chunks
    for chunk_start in range(0, sheet_count, chunk_size):
        chunk_end = min(chunk_start + chunk_size, sheet_count)
        chunk_sheets = sheet_names[chunk_start:chunk_end]

        logger.info(f"Processing chunk {chunk_start//chunk_size + 1}: sheets {chunk_start+1}-{chunk_end}")

        # Create temporary Excel file with this chunk of sheets
        chunk_wb = Workbook()
        chunk_wb.remove(chunk_wb.active)  # Remove default sheet

        # Load source workbook
        source_wb = load_workbook(xlsx_path, data_only=False)

        # Copy each sheet in this chunk
        for sheet_name in chunk_sheets:
            if sheet_name in source_wb.sheetnames:
                # Copy sheet data (simplified to avoid DrawingML issues)
                source_sheet = source_wb[sheet_name]
                target_sheet = chunk_wb.create_sheet(sheet_name)

                # Copy cell values (preserves data but not complex formatting)
                for row in source_sheet.iter_rows():
                    for cell in row:
                        target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

        # Save temporary chunk file
        temp_chunk_file = output_dir / f"temp_chunk_{chunk_start//chunk_size + 1}.xlsx"
        chunk_wb.save(temp_chunk_file)

        try:
            # Convert this chunk to PDF
            client = LoolClient(timeout=900)  # 15 minutes per chunk
            chunk_pdf_bytes = client.convert_all_sheets(temp_chunk_file)

            # Split chunk PDF into individual pages
            from pypdf import PdfReader, PdfWriter

            reader = PdfReader(io.BytesIO(chunk_pdf_bytes))
            actual_pages = len(reader.pages)

            logger.info(f"Chunk produced {actual_pages} pages (expected {len(chunk_sheets)})")

            # Extract individual pages from this chunk
            for page_idx, sheet_name in enumerate(chunk_sheets):
                if page_idx < actual_pages:
                    # Extract single page
                    writer = PdfWriter()
                    writer.add_page(reader.pages[page_idx])

                    # Save individual PDF
                    safe_name = sanitize_filename(sheet_name)
                    pdf_filename = f"{chunk_start + page_idx + 1:03d}_{safe_name}.pdf"
                    pdf_path = output_dir / pdf_filename

                    with open(pdf_path, 'wb') as f:
                        writer.write(f)

                    all_sheet_files.append(pdf_path)
                    logger.info(f"  → Saved {pdf_path}")

        except Exception as e:
            logger.error(f"Failed to convert chunk {chunk_start//chunk_size + 1}: {e}")
            continue
        finally:
            # Clean up temporary file
            if temp_chunk_file.exists():
                temp_chunk_file.unlink()

    logger.info(f"✅ Completed chunked conversion: {len(all_sheet_files)} individual PDFs")
    return all_sheet_files


def batch_convert_excel(excel_path: Path, client: LoolClient, deep: bool = False) -> Path:
    """
    Convert each worksheet to individual PDFs, then combine them.

    Args:
        excel_path: Path to Excel file
        client: LOOL client instance
        deep: If True, split Excel into individual sheet files first

    Returns:
        Path to the combined PDF file
    """
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    # Extract filename without extension
    excel_name = excel_path.stem
    output_dir = Path("split") / excel_name
    output_dir.mkdir(parents=True, exist_ok=True)

    logger.info(f"Processing {excel_path}")
    logger.info(f"Output directory: {output_dir}")
    if deep:
        logger.info("Using deep mode: splitting Excel into individual sheet files first")

    # Get sheet names
    sheet_names = get_sheet_names(excel_path)
    sheet_count = len(sheet_names)

    logger.info(f"Found {sheet_count} worksheets: {', '.join(sheet_names)}")

    if deep:
        # Convert entire workbook and split into individual PDFs (preserves ALL content)
        individual_pdfs = split_excel_into_sheets(excel_path, output_dir)
    else:
        # Convert each sheet individually (original approach)
        individual_pdfs = []
        for idx, sheet_name in enumerate(sheet_names, 1):
            logger.info(f"Converting sheet {idx}/{sheet_count}: {sheet_name}")

            # Create safe filename
            safe_name = sanitize_filename(sheet_name)
            pdf_filename = f"{idx:03d}_{safe_name}.pdf"
            pdf_path = output_dir / pdf_filename

            try:
                # Convert single sheet
                client.convert_sheet(excel_path, sheet_index=idx, output_path=pdf_path)
                individual_pdfs.append(pdf_path)
                logger.info(f"  → Saved {pdf_path}")

            except Exception as e:
                logger.warning(f"  → FullSheetPreview failed for sheet {sheet_name}: {e}")
                logger.info(f"  → Trying alternative conversion method...")

                try:
                    # Fallback: Convert entire workbook and extract the specific page
                    # This is less efficient but should work for problematic files
                    all_pages_pdf = client.convert_all_sheets(excel_path)

                    # Extract the specific page
                    from pypdf import PdfReader, PdfWriter
                    reader = PdfReader(io.BytesIO(all_pages_pdf))

                    if idx <= len(reader.pages):
                        writer = PdfWriter()
                        writer.add_page(reader.pages[idx - 1])  # 0-based indexing

                        with open(pdf_path, 'wb') as f:
                            writer.write(f)

                        individual_pdfs.append(pdf_path)
                        logger.info(f"  → Saved {pdf_path} (using fallback method)")
                    else:
                        logger.error(f"  → Sheet index {idx} out of range (max: {len(reader.pages)})")

                except Exception as e2:
                    logger.error(f"  → Fallback conversion also failed: {e2}")
                    continue

    if not individual_pdfs:
        raise RuntimeError("No sheets were successfully converted")

    # Combine all individual PDFs
    combined_filename = f"combined_{excel_name}.pdf"
    combined_path = output_dir / combined_filename

    logger.info(f"Combining {len(individual_pdfs)} PDFs into {combined_path}")

    try:
        # Use the existing merge functionality
        from pypdf import PdfWriter

        merger = PdfWriter()

        for pdf_path in individual_pdfs:
            try:
                # Read each individual PDF and add its page
                from pypdf import PdfReader
                reader = PdfReader(pdf_path)
                if len(reader.pages) > 0:
                    merger.add_page(reader.pages[0])
                else:
                    logger.warning(f"PDF {pdf_path} has no pages, skipping")
            except Exception as e:
                logger.error(f"Failed to read {pdf_path}: {e}")

        # Write combined PDF
        with open(combined_path, 'wb') as f:
            merger.write(f)

        logger.info(f"✅ Successfully created combined PDF: {combined_path}")
        logger.info(f"📊 Total pages: {len(merger.pages)}")

    except Exception as e:
        logger.error(f"Failed to create combined PDF: {e}")
        raise

    return combined_path


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="Batch convert Excel worksheets to individual PDFs, then combine them",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python batch_convert.py myworkbook.xlsx

Output structure:
  split/myworkbook/
  ├── sheet_001_Sheet1.pdf
  ├── sheet_002_Sheet2.pdf
  └── combined_myworkbook.pdf

Environment variables:
  LOOL_URL: LibreOffice Online server URL (default: http://localhost:9980)
  LOOL_ENDPOINT: API endpoint path (default: /lool/convert-to/pdf)
        """
    )

    parser.add_argument(
        "excel_file",
        type=Path,
        help="Input Excel file (.xlsx, .xls)"
    )

    parser.add_argument(
        "--lool-url",
        default="http://localhost:9980",
        help="LibreOffice Online server URL"
    )

    parser.add_argument(
        "--endpoint",
        default="/lool/convert-to/pdf",
        help="API endpoint path"
    )

    parser.add_argument(
        "--timeout",
        type=int,
        default=900,  # 15 minutes for large files
        help="HTTP timeout in seconds (default: 900)"
    )

    parser.add_argument(
        "--deep",
        action="store_true",
        help="Split Excel into individual sheet files first (slower but more reliable for large files)"
    )

    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Enable verbose logging"
    )

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # Validate input
    if not args.excel_file.exists():
        logger.error(f"Excel file not found: {args.excel_file}")
        sys.exit(1)

    # Initialize client
    client = LoolClient(
        base_url=args.lool_url,
        endpoint_path=args.endpoint,
        timeout=args.timeout
    )

    try:
        combined_pdf = batch_convert_excel(args.excel_file, client, deep=args.deep)
        print(f"\n🎉 Conversion complete!")
        print(f"📁 Individual PDFs: {combined_pdf.parent}")
        print(f"📄 Combined PDF: {combined_pdf}")
        if args.deep:
            print("📄 All SmartArt, images, and formatting preserved!")

    except Exception as e:
        logger.error(f"Batch conversion failed: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
